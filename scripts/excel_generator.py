import os
import datetime
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from scripts.database import Job
from scripts.logger import log_info, log_error

REPORTS_DIR = os.getenv("REPORTS_DIR", "/app/reports")
if not os.path.exists(REPORTS_DIR):
    os.makedirs(REPORTS_DIR, exist_ok=True)

def generate_daily_excel(db: Session) -> str:
    """
    Queries all jobs scraped in the last 24 hours, creates a highly styled, professional
    Excel sheet and returns the file path.
    """
    try:
        # Fetch jobs created in the last 24 hours
        time_threshold = datetime.datetime.utcnow() - datetime.timedelta(hours=24)
        jobs = db.query(Job).filter(Job.created_at >= time_threshold).order_by(Job.ats_score.desc().nullslast()).all()
        
        if not jobs:
            log_info("No jobs scraped in the last 24 hours. Excel report will be created but empty.", "ExcelGenerator")

        wb = Workbook()
        ws = wb.active
        ws.title = "Scraped Jobs"
        
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
        
        # Freeze panes below header
        ws.freeze_panes = "A2"

        # Headers
        headers = [
            "Company", "Job Title", "Location", "Experience", "Skills Match", 
            "ATS Score", "Priority", "Posted Date", "Platform", "Apply Link", 
            "Apply Email / Contact", "Status", "Date Applied", "Notes", 
            "Follow-up Date", "Recruiter", "Recruiter LinkedIn"
        ]
        
        # Styles
        font_name = "Segoe UI"
        header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Navy
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Fills for priority
        fill_high = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
        font_high = Font(name=font_name, size=10, color="006100", bold=True)
        
        fill_medium = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Light Yellow
        font_medium = Font(name=font_name, size=10, color="9C6500", bold=True)
        
        fill_low = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid") # Light Grey
        font_low = Font(name=font_name, size=10, color="595959")

        # Fills for ATS Scores
        fill_ats_good = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid") # Soft green
        font_ats_good = Font(name=font_name, size=10, color="274E13", bold=True)
        
        fill_ats_bad = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid") # Soft red
        font_ats_bad = Font(name=font_name, size=10, color="660000")

        # Write header
        ws.row_dimensions[1].height = 28
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write data rows
        current_row = 2
        for job in jobs:
            ws.row_dimensions[current_row].height = 20
            
            # Simple fields
            ws.cell(row=current_row, column=1, value=job.company_name)
            ws.cell(row=current_row, column=2, value=job.title)
            ws.cell(row=current_row, column=3, value=job.location)
            ws.cell(row=current_row, column=4, value=job.experience)
            ws.cell(row=current_row, column=5, value=job.skills or "")
            
            # ATS score with custom conditional styling
            ats_cell = ws.cell(row=current_row, column=6, value=job.ats_score)
            if job.ats_score is not None:
                if job.ats_score >= 80:
                    ats_cell.fill = fill_ats_good
                    ats_cell.font = font_ats_good
                elif job.ats_score < 50:
                    ats_cell.fill = fill_ats_bad
                    ats_cell.font = font_ats_bad
            
            # Priority with custom styling
            prio_cell = ws.cell(row=current_row, column=7, value=job.priority or "Low")
            if job.priority == "High":
                prio_cell.fill = fill_high
                prio_cell.font = font_high
            elif job.priority == "Medium":
                prio_cell.fill = fill_medium
                prio_cell.font = font_medium
            else:
                prio_cell.fill = fill_low
                prio_cell.font = font_low
                
            prio_cell.alignment = Alignment(horizontal="center")
                
            # Date
            posted_date_str = job.posted_date.strftime("%Y-%m-%d") if job.posted_date else ""
            ws.cell(row=current_row, column=8, value=posted_date_str)
            ws.cell(row=current_row, column=9, value=job.platform)
            
            # Apply URL as blue underlined Hyperlink
            link_cell = ws.cell(row=current_row, column=10, value="Apply Link")
            if job.apply_url:
                link_cell.hyperlink = job.apply_url
                link_cell.font = Font(name=font_name, size=10, color="0563C1", underline="single")
            
            # Recruiter Email
            email_cell = ws.cell(row=current_row, column=11, value=job.recruiter_email or "")
            if job.recruiter_email and "@" in job.recruiter_email:
                email_cell.hyperlink = f"mailto:{job.recruiter_email}"
                email_cell.font = Font(name=font_name, size=10, color="0563C1", underline="single")

            # Status tracking (manual fields)
            ws.cell(row=current_row, column=12, value=job.status or "Not Applied")
            
            applied_date_str = job.date_applied.strftime("%Y-%m-%d") if job.date_applied else ""
            ws.cell(row=current_row, column=13, value=applied_date_str)
            ws.cell(row=current_row, column=14, value=job.notes or "")
            
            follow_up_str = job.follow_up_date.strftime("%Y-%m-%d") if job.follow_up_date else ""
            ws.cell(row=current_row, column=15, value=follow_up_str)
            ws.cell(row=current_row, column=16, value="")
            ws.cell(row=current_row, column=17, value="")
            
            # Apply fonts & alignment & borders to all standard cells in the row
            for col_idx in range(1, 18):
                cell = ws.cell(row=current_row, column=col_idx)
                if col_idx not in [6, 7, 10, 11]: # Skip already customized cells (ATS, Priority, Links)
                    cell.font = Font(name=font_name, size=10)
                cell.border = thin_border
                
                # Center align status, dates, platform
                if col_idx in [3, 4, 8, 9, 12, 13, 15]:
                    cell.alignment = Alignment(horizontal="center")
            
            current_row += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.hyperlink:
                    val = "Apply Link" # Cap the link column width calculation
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Apply auto-filter across headers
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{current_row - 1}"

        # Write to file
        today_str = datetime.date.today().strftime("%Y_%m_%d")
        filename = f"Jobs_{today_str}.xlsx"
        full_path = os.path.join(REPORTS_DIR, filename)
        wb.save(full_path)
        log_info(f"Excel tracker generated successfully at: {full_path}", "ExcelGenerator")
        return full_path

    except Exception as e:
        log_error(f"Failed to generate Excel report: {e}", "ExcelGenerator")
        raise e
